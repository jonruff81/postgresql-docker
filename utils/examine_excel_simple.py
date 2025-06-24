#!/usr/bin/env python3
"""
Simple Excel file examiner without pandas dependency
"""

import os
from openpyxl import load_workbook

def examine_excel_file(filepath):
    """Examine a single Excel file"""
    try:
        filename = os.path.basename(filepath)
        print(f"\n=== EXAMINING: {filename} ===")
        
        # Load workbook
        wb = load_workbook(filepath, read_only=True)
        print(f"Worksheets: {wb.sheetnames}")
        
        # Examine first worksheet
        ws = wb.active
        print(f"Active worksheet: {ws.title}")
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Dimensions: {max_row} rows x {max_col} columns")
        
        # Get ALL column headers
        headers = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Col{col}")
        
        print(f"\nALL {len(headers)} column headers:")
        for i, header in enumerate(headers, 1):
            print(f"  {i:2d}: {header}")
        
        # Show first data row with ALL columns
        print(f"\nFirst data row (Row 2) - ALL COLUMNS:")
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=2, column=col).value
            header = headers[col-1]
            print(f"  {header}: {cell_value}")
        
        # Look for key pricing/quantity columns
        pricing_cols = ['QTY', 'UoM', 'Factor', 'Price', 'ExtendedPrice', 'Vendor']
        print(f"\nKey pricing data from Row 2:")
        for col_name in pricing_cols:
            if col_name in headers:
                col_index = headers.index(col_name) + 1
                cell_value = ws.cell(row=2, column=col_index).value
                print(f"  {col_name}: {cell_value}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error examining {filepath}: {e}")
        return False

def main():
    """Examine all Excel files"""
    directory = 'PlanElevOptions'
    
    if not os.path.exists(directory):
        print(f"Directory {directory} does not exist")
        return
    
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    print(f"Found {len(files)} Excel files")
    
    # Examine just the first file for now
    if files:
        filepath = os.path.join(directory, files[0])
        examine_excel_file(filepath)
    else:
        print("No Excel files found")

if __name__ == "__main__":
    main() 