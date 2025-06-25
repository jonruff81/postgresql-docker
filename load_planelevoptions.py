#!/usr/bin/env python3
"""
Simple data loader for PlanElevOptions Excel files
Focuses on loading data directly without numpy dependency issues
"""

import os
import sys
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database connection settings (matching docker-compose.yml)
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'takeoff_pricing_db',
    'user': 'Jon',
    'password': 'Transplant4real'
}

def connect_db():
    """Connect to PostgreSQL database"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        logger.error(f"Database connection failed: {e}")
        return None

def parse_filename(filename):
    """Parse plan info from filename"""
    # Format: PlanName_Elevation_Foundation_Type.xlsx
    # Example: Barringer_A_Crawl_Base Home.xlsx
    parts = filename.replace('.xlsx', '').split('_')
    
    if len(parts) >= 3:
        plan_name = parts[0]
        elevation = parts[1]
        foundation = parts[2]
        option_type = ' '.join(parts[3:]) if len(parts) > 3 else 'Standard'
        
        return {
            'plan_name': plan_name,
            'elevation': elevation,
            'foundation': foundation,
            'option_type': option_type
        }
    return None

def load_excel_data(file_path):
    """Load data from Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Look for data in different sheets
        data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find square footage data
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).lower()
                        
                        # Look for square footage indicators
                        if 'heated' in cell_text and 'sf' in cell_text:
                            # Try to find the value in adjacent cells
                            for offset in [1, 2, -1, -2]:
                                try:
                                    adj_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if adj_cell.value and isinstance(adj_cell.value, (int, float)):
                                        if 'inside' in cell_text:
                                            data['heated_sf_inside_studs'] = int(adj_cell.value)
                                        elif 'outside' in cell_text:
                                            data['heated_sf_outside_studs'] = int(adj_cell.value)
                                        break
                                except:
                                    continue
                        
                        # Look for bedroom/bathroom counts
                        if 'bedroom' in cell_text or 'br' in cell_text:
                            for offset in [1, 2, -1, -2]:
                                try:
                                    adj_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if adj_cell.value and isinstance(adj_cell.value, (int, float)):
                                        data['bedroom_count'] = int(adj_cell.value)
                                        break
                                except:
                                    continue
                        
                        if 'bathroom' in cell_text or 'bath' in cell_text:
                            for offset in [1, 2, -1, -2]:
                                try:
                                    adj_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                    if adj_cell.value and isinstance(adj_cell.value, (int, float)):
                                        data['bathroom_count'] = int(adj_cell.value)
                                        break
                                except:
                                    continue
        
        return data
    except Exception as e:
        logger.error(f"Error loading Excel file {file_path}: {e}")
        return {}

def ensure_plan(conn, plan_name):
    """Ensure plan record exists and return its ID"""
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Check if plan exists
        cursor.execute("""
            SELECT plan_id FROM takeoff.plans WHERE plan_name = %s
        """, (plan_name,))
        
        result = cursor.fetchone()
        if result:
            return result['plan_id']
        
        # Create new plan
        cursor.execute("""
            INSERT INTO takeoff.plans (plan_name) VALUES (%s) RETURNING plan_id
        """, (plan_name,))
        
        result = cursor.fetchone()
        conn.commit()
        return result['plan_id']
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error ensuring plan: {e}")
        return None
    finally:
        cursor.close()

def ensure_plan_elevation(conn, plan_id, plan_info):
    """Ensure plan_elevation record exists and return its ID"""
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Check if plan_elevation exists
        cursor.execute("""
            SELECT plan_elevation_id FROM takeoff.plan_elevations 
            WHERE plan_id = %s AND elevation_name = %s AND foundation = %s
        """, (plan_id, plan_info['elevation'], plan_info['foundation']))
        
        result = cursor.fetchone()
        if result:
            return result['plan_elevation_id']
        
        # Create new plan_elevation
        cursor.execute("""
            INSERT INTO takeoff.plan_elevations (plan_id, elevation_name, foundation, plan_full_name)
            VALUES (%s, %s, %s, %s)
            RETURNING plan_elevation_id
        """, (plan_id, plan_info['elevation'], plan_info['foundation'], 
              f"{plan_info['plan_name']} {plan_info['elevation']} {plan_info['foundation']}"))
        
        result = cursor.fetchone()
        conn.commit()
        return result['plan_elevation_id']
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error ensuring plan_elevation: {e}")
        return None
    finally:
        cursor.close()

def load_plan_option(conn, plan_elevation_id, plan_info, excel_data):
    """Load plan option data"""
    cursor = conn.cursor()
    
    try:
        # Check if option already exists
        cursor.execute("""
            SELECT plan_option_id FROM takeoff.plan_options 
            WHERE plan_elevation_id = %s AND option_name = %s
        """, (plan_elevation_id, plan_info['option_type']))
        
        existing = cursor.fetchone()
        if existing:
            logger.info(f"Option {plan_info['option_type']} already exists for {plan_info['plan_name']}_{plan_info['elevation']}")
            return existing[0]
        
        # Insert new option
        cursor.execute("""
            INSERT INTO takeoff.plan_options (
                plan_elevation_id, option_name, option_type, option_description,
                bedroom_count, bathroom_count,
                heated_sf_inside_studs, heated_sf_outside_studs,
                total_sf_inside_studs, total_sf_outside_studs
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING plan_option_id
        """, (
            plan_elevation_id,
            plan_info['option_type'],
            plan_info['foundation'],
            f"{plan_info['plan_name']} {plan_info['elevation']} {plan_info['foundation']} {plan_info['option_type']}",
            excel_data.get('bedroom_count'),
            excel_data.get('bathroom_count'),
            excel_data.get('heated_sf_inside_studs'),
            excel_data.get('heated_sf_outside_studs'),
            excel_data.get('heated_sf_inside_studs'),  # Default total to heated
            excel_data.get('heated_sf_outside_studs')   # Default total to heated
        ))
        
        result = cursor.fetchone()
        conn.commit()
        logger.info(f"Loaded option: {plan_info['plan_name']}_{plan_info['elevation']}_{plan_info['foundation']}_{plan_info['option_type']}")
        return result[0]
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error loading plan option: {e}")
        return None
    finally:
        cursor.close()

def main():
    """Main data loading function"""
    planelevOptions_dir = "PlanElevOptions"
    
    if not os.path.exists(planelevOptions_dir):
        logger.error(f"Directory {planelevOptions_dir} not found")
        return
    
    # Connect to database
    conn = connect_db()
    if not conn:
        logger.error("Failed to connect to database")
        return
    
    try:
        # Get list of Excel files
        excel_files = [f for f in os.listdir(planelevOptions_dir) if f.endswith('.xlsx')]
        logger.info(f"Found {len(excel_files)} Excel files to process")
        
        for excel_file in excel_files:
            logger.info(f"Processing: {excel_file}")
            
            # Parse filename
            plan_info = parse_filename(excel_file)
            if not plan_info:
                logger.warning(f"Could not parse filename: {excel_file}")
                continue
            
            # Load Excel data
            file_path = os.path.join(planelevOptions_dir, excel_file)
            excel_data = load_excel_data(file_path)
            
            # Ensure plan exists
            plan_id = ensure_plan(conn, plan_info['plan_name'])
            if not plan_id:
                logger.error(f"Failed to ensure plan for {excel_file}")
                continue
            
            # Ensure plan_elevation exists
            plan_elevation_id = ensure_plan_elevation(conn, plan_id, plan_info)
            if not plan_elevation_id:
                logger.error(f"Failed to ensure plan_elevation for {excel_file}")
                continue
            
            # Load plan option
            plan_option_id = load_plan_option(conn, plan_elevation_id, plan_info, excel_data)
            if plan_option_id:
                logger.info(f"Successfully loaded plan option ID: {plan_option_id}")
        
        # Show summary
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT COUNT(*) as count FROM takeoff.plan_elevations")
        pe_count = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM takeoff.plan_options")
        po_count = cursor.fetchone()['count']
        
        logger.info(f"Summary: {pe_count} plan elevations, {po_count} plan options")
        cursor.close()
        
    finally:
        conn.close()

if __name__ == "__main__":
    main()
