#!/usr/bin/env python3
"""
Simple data loader using openpyxl directly to avoid NumPy conflicts
"""

import os
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Database connection settings
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'takeoff_pricing_db',
    'user': 'Jon',
    'password': 'Transplant4real'
}

def safe_int(value):
    """Safely convert to int"""
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except:
        return None

def safe_float(value):
    """Safely convert to float"""
    if value is None or value == '':
        return None
    try:
        return float(value)
    except:
        return None

def safe_date(value):
    """Safely convert to date"""
    if value is None or value == '':
        return None
    try:
        if isinstance(value, str):
            # Parse date string like "4.7.25"
            if '.' in value:
                parts = value.split('.')
                if len(parts) == 3:
                    month, day, year = parts
                    year = f"20{year}" if len(year) == 2 else year
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        return str(value)
    except:
        return None

def get_or_create(cursor, table, where_data, insert_data):
    """Get existing record or create new one based on a dictionary of where conditions."""
    where_clause = " AND ".join([f"{key} = %s" for key in where_data.keys()])
    query = f"SELECT * FROM {table} WHERE {where_clause}"
    cursor.execute(query, list(where_data.values()))
    result = cursor.fetchone()
    
    # Handle irregular plurals
    id_col_map = {
        'divisions': 'division_id',
        'communities': 'community_id',
        'plans': 'plan_id',
        'plan_elevations': 'plan_elevation_id',
        'plan_options': 'plan_option_id',
        'cost_groups': 'cost_group_id',
        'cost_codes': 'cost_code_id',
        'vendors': 'vendor_id',
        'items': 'item_id',
        'products': 'product_id'
    }
    id_col = id_col_map[table]
    
    if result:
        return result[id_col]
    else:
        # Combine where_data and insert_data for the insert statement
        final_insert_data = where_data.copy()
        final_insert_data.update(insert_data)

        cols = ', '.join(final_insert_data.keys())
        placeholders = ', '.join(['%s'] * len(final_insert_data))
        cursor.execute(f"""
            INSERT INTO {table} ({cols}) VALUES ({placeholders})
            RETURNING {id_col}
        """, list(final_insert_data.values()))
        return cursor.fetchone()[id_col]

def load_excel_file(filepath):
    """Load one Excel file into the database"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SET search_path TO takeoff, public")
        
        filename = os.path.basename(filepath)
        logger.info(f"Loading file: {filename}")
        
        # Parse filename: PlanName_Elevation_Foundation_OptionType.xlsx
        # Parse filename: PlanName_Elevation_Foundation_OptionType.xlsx
        # Parse filename: PlanName_Elevation_Foundation_OptionType.xlsx
        clean_name = filename.replace('.xlsx', '').strip()
        # Normalize spaces to underscores for consistent splitting
        normalized_name = clean_name.replace(' ', '_')
        parts = [part for part in normalized_name.split('_') if part] # remove empty parts
        plan_name = parts[0]
        elevation_name = parts[1]
        foundation = parts[2]
        # Re-join the remaining parts to form the option name
        option_name = ' '.join(parts[3:])
        
        # Load Excel file
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        if not ws:
            logger.error(f"No active worksheet found in {filepath}")
            return False
        
        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Col{col}")
        
        logger.info(f"Found {len(headers)} columns, {ws.max_row-1} data rows")
        
        # Get first row for plan-level data
        first_row_data = {}
        for col, header in enumerate(headers, 1):
            first_row_data[header] = ws.cell(row=2, column=col).value
        
        # Create plan structure
        plan_id = get_or_create(cursor, 'plans', {'plan_name': plan_name}, {
            'plan_name': plan_name,
            'architect': first_row_data.get('Architect'),
            'engineer': first_row_data.get('Engineer')
        })
        
        plan_elevation_id = get_or_create(cursor, 'plan_elevations',
            {'plan_id': plan_id, 'elevation_name': elevation_name, 'foundation': foundation},
            {
                'heated_sf_inside_studs': safe_int(first_row_data.get('HeatedSF(InsideStuds)')),
                'total_sf_outside_studs': safe_int(first_row_data.get('TotalSF(OutsideStuds)'))
            }
        )
        
        plan_option_id = get_or_create(cursor, 'plan_options', {'plan_elevation_id': plan_elevation_id, 'option_name': option_name}, {
            'plan_elevation_id': plan_elevation_id,
            'option_name': option_name,
            'option_type': first_row_data.get('OptionType', option_name)
        })
        
        
        # Create job (template)
        # Check for existing template job or create a new one
        cursor.execute("SELECT job_id FROM jobs WHERE plan_option_id = %s AND is_template = TRUE", (plan_option_id,))
        existing_job = cursor.fetchone()
        
        if existing_job:
            job_id = existing_job['job_id']
            logger.info(f"Found existing template job {job_id} for plan_option_id {plan_option_id}")
        else:
            job_name = f"{plan_name} {elevation_name} {foundation} {option_name} Template"
            cursor.execute("""
                INSERT INTO jobs (
                    job_name, plan_option_id, is_template
                ) VALUES (%s, %s, %s) RETURNING job_id
            """, (
                job_name,
                plan_option_id,
                True  # This is a template
            ))
            job_id = cursor.fetchone()['job_id']
            logger.info(f"Created new template job {job_id} for plan_option_id {plan_option_id}")
        
        # Process each takeoff line
        processed_rows = 0
        for row_num in range(2, ws.max_row + 1):
            try:
                # Get row data
                row_data = {}
                for col, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row=row_num, column=col).value
                
                # Skip rows with missing critical data
                if not row_data.get('Item') or not row_data.get('TakeoffID'):
                    continue
                
                # Create cost structure
                cost_group_id = get_or_create(cursor, 'cost_groups', {'cost_group_code': row_data.get('CostGroup')}, {
                    'cost_group_code': row_data.get('CostGroup', ''),
                    'cost_group_name': row_data.get('CostGroupName', '')
                })
                
                cost_code_id = get_or_create(cursor, 'cost_codes', {'cost_code': row_data.get('CostCodeNumber')}, {
                    'cost_group_id': cost_group_id,
                    'cost_code': row_data.get('CostCodeNumber', ''),
                    'cost_code_description': row_data.get('CostCodeDescription', '')
                })
                
                # Create item
                # Get item and its potential formula
                cursor.execute("""
                    SELECT i.item_id, f.formula_text, f.depends_on_fields
                    FROM items i
                    LEFT JOIN formulas f ON i.formula_id = f.formula_id
                    WHERE i.item_name = %s
                """, (row_data.get('Item'),))
                item_details = cursor.fetchone()

                if item_details:
                    item_id = item_details['item_id']
                    formula_text = item_details['formula_text']
                    depends_on = item_details['depends_on_fields']
                else:
                    # Create item if it doesn't exist
                    item_id = get_or_create(cursor, 'items', {'item_name': row_data.get('Item')}, {
                        'cost_code_id': cost_code_id,
                        'qty_type': row_data.get('QtyType'),
                        'takeoff_type': row_data.get('TakeoffType'),
                        'default_unit': row_data.get('UoM')
                    })
                    formula_text = None
                    depends_on = None
                
                # Create product
                product_id = get_or_create(cursor, 'products', {'item_id': item_id, 'product_description': row_data.get('ItemDescription', '')}, {
                    'item_id': item_id,
                    'product_description': row_data.get('ItemDescription', ''),
                    'model': row_data.get('Model')
                })
                
                # Create vendor if exists
                vendor_id = None
                if row_data.get('Vendor'):
                    vendor_id = get_or_create(cursor, 'vendors', {'vendor_name': row_data.get('Vendor')}, {
                        'vendor_name': row_data.get('Vendor')
                    })
                
                # Insert takeoff line
                # Calculate quantity
                quantity = safe_float(row_data.get('QTY'))
                quantity_source = 'Excel'

                if formula_text:
                    try:
                        # Build a context for the formula evaluation
                        eval_context = {
                            'heated_sf_inside_studs': safe_float(first_row_data.get('HeatedSF(InsideStuds)')),
                            'total_sf_outside_studs': safe_float(first_row_data.get('TotalSF(OutsideStuds)'))
                            # Add other fields from depends_on as needed
                        }
                        # Basic, unsafe eval. For production, use a safe evaluation library.
                        quantity = eval(formula_text, {"__builtins__": {}}, eval_context)
                        quantity_source = f"Formula: {formula_text}"
                    except Exception as e:
                        logger.warning(f"Could not evaluate formula '{formula_text}' for item {row_data.get('Item')}: {e}")
                

                # Insert takeoff line
                cursor.execute("""
                    INSERT INTO takeoffs (
                        takeoff_id, job_id, product_id, vendor_id,
                        quantity, unit_price, price_factor, extended_price,
                        unit_of_measure, notes,
                        room_name, floor_level, is_heated,
                        spec_name, spec_description, quantity_source
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (takeoff_id) DO UPDATE SET
                        quantity = EXCLUDED.quantity,
                        unit_price = EXCLUDED.unit_price,
                        extended_price = EXCLUDED.extended_price,
                        quantity_source = EXCLUDED.quantity_source
                """, (
                    safe_int(row_data.get('TakeoffID')),
                    job_id,
                    product_id,
                    vendor_id,
                    quantity,
                    safe_float(row_data.get('Price')),
                    safe_float(row_data.get('Factor', 1.0)),
                    safe_float(row_data.get('ExtendedPrice')),
                    row_data.get('UoM'),
                    row_data.get('Notes'),
                    row_data.get('Room'),
                    row_data.get('FloorLevel'),
                    row_data.get('IsHeated') if row_data.get('IsHeated') is not None else None,
                    row_data.get('SpecName'),
                    row_data.get('SpecDescription'),
                    quantity_source
                ))
                
                processed_rows += 1
                
            except psycopg2.OperationalError as op_err:
                logger.error(f"Connection error on row {row_num}: {op_err}")
                # The connection is likely dead, so we can't continue with this file.
                # Return False to signal failure for this file.
                return False
            except psycopg2.Error as db_err:
                logger.error(f"Database error on row {row_num}: {db_err}")
                conn.rollback()  # Rollback the transaction on other DB errors
            except Exception as e:
                logger.warning(f"Error processing row {row_num}: {e}")
                continue
        
        conn.commit()
        logger.info(f"Successfully loaded {processed_rows} rows from {filename}")
        return True
        
    except Exception as e:
        logger.error(f"Error loading file {filepath}: {e}")
        return False
    finally:
        # Ensure resources are always released
        if 'wb' in locals() and wb:
            wb.close()
        if 'cursor' in locals() and not cursor.closed:
            cursor.close()
        if 'conn' in locals() and not conn.closed:
            conn.close()

def main():
    """Load first three Excel files"""
    directory = 'PlanElevOptions'
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    
    if not files:
        logger.error("No Excel files found")
        return

    loaded_count = 0
    for i, filename in enumerate(files[:3]):
        filepath = os.path.join(directory, filename)
        success = load_excel_file(filepath)
        if success:
            loaded_count += 1
            logger.info(f"File {i+1} of 3 loaded successfully!")
        else:
            logger.error(f"File {i+1} of 3 loading failed")
            
    logger.info(f"Finished loading. {loaded_count} of 3 files were processed successfully.")

if __name__ == "__main__":
    main() 